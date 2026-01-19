"""
NVH Task Views
"""
import os
import uuid
import shutil
from datetime import datetime, timedelta
from io import BytesIO
from django.conf import settings
from django.db import transaction
from django.db.models import Q
from django.db.models.functions import TruncDate
from django.core.exceptions import ValidationError as DjangoValidationError
from django.utils import timezone
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from utils.response import Response
from .models import (
    MainRecord, EntryExit, TestInfo, DocApproval,
    TestProcessAttachment, TestProcessList, STATUS_DRAFT
)
from .serializers import (
    MainRecordListSerializer, MainRecordDetailSerializer, MainRecordCreateUpdateSerializer,
    EntryExitSerializer, TestInfoSerializer, DocApprovalSerializer,
    TestProcessAttachmentSerializer, TestProcessListSerializer
)
from . import services


# ==================== 文件上传常量 ====================

ALLOWED_IMAGE_TYPES = ['image/jpeg', 'image/png', 'image/webp', 'image/gif']
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
TEMP_UPLOAD_DIR = 'nvh_task/_temp'  # 临时目录前缀


# ==================== 文件上传辅助函数 ====================

def get_final_upload_path(temp_path: str, upload_type: str) -> str:
    """
    将临时路径转换为最终路径（不做移动）
    :param temp_path: 临时文件相对路径（如 nvh_task/_temp/xxx.jpg）
    :param upload_type: 上传类型
    :return: 最终文件相对路径
    """
    if not temp_path:
        return ''

    # 如果已经是最终路径（不在临时目录），直接返回
    if not temp_path.startswith(TEMP_UPLOAD_DIR):
        return temp_path

    type_to_dir = {
        'teardown_record': 'nvh_task/teardown_record',
        'nvh_test_process': 'nvh_task/nvh_test_process',
        'nvh_task_approval': 'nvh_task/nvh_task_approval',
    }
    final_dir = type_to_dir.get(upload_type, 'nvh_task/nvh_test_process')
    filename = os.path.basename(temp_path)
    return f"{final_dir}/{filename}"


def confirm_file_upload(temp_path: str, upload_type: str) -> str:
    """
    确认文件上传：将临时文件移动到最终目录
    :param temp_path: 临时文件相对路径（如 nvh_task/_temp/xxx.jpg）
    :param upload_type: 上传类型
    :return: 最终文件相对路径
    """
    if not temp_path:
        return ''

    # 如果已经是最终路径（不在临时目录），直接返回
    if not temp_path.startswith(TEMP_UPLOAD_DIR):
        return temp_path

    final_relative_path = get_final_upload_path(temp_path, upload_type)
    final_dir = os.path.dirname(final_relative_path)

    src_path = os.path.join(settings.MEDIA_ROOT, temp_path)
    dest_dir = os.path.join(settings.MEDIA_ROOT, final_dir)
    os.makedirs(dest_dir, exist_ok=True)
    dest_path = os.path.join(settings.MEDIA_ROOT, final_relative_path)

    # 移动文件
    if os.path.exists(src_path):
        shutil.move(src_path, dest_path)

    return final_relative_path


def cleanup_temp_file(temp_path: str):
    """清理临时文件"""
    if not temp_path or not temp_path.startswith(TEMP_UPLOAD_DIR):
        return
    full_path = os.path.join(settings.MEDIA_ROOT, temp_path)
    if os.path.exists(full_path):
        try:
            os.remove(full_path)
        except Exception:
            pass


# ==================== 分页工具 ====================

def get_pagination_params(request):
    """获取分页参数"""
    try:
        page = int(request.GET.get('page', '1'))
    except (ValueError, TypeError):
        page = 1

    default_page_size = getattr(settings, 'REST_FRAMEWORK', {}).get('PAGE_SIZE', 20)
    try:
        page_size = int(request.GET.get('page_size', str(default_page_size)))
    except (ValueError, TypeError):
        page_size = default_page_size

    return max(1, page), max(1, min(page_size, 100))


def paginate_queryset(queryset, page, page_size):
    """分页查询集"""
    total = queryset.count()
    start = (page - 1) * page_size
    end = start + page_size
    items = queryset[start:end]
    return items, total


# ==================== MainRecord 视图 ====================

@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def main_record_list(request):
    """主记录列表（含筛选）/ 创建"""
    if request.method == 'GET':
        # 排序：按测试人员聚合，组内按排期开始时间倒序，同时间按id倒序
        queryset = MainRecord.objects.select_related(
            'entry_exit', 'test_info', 'doc_approval'
        ).order_by('-schedule_start', 'tester_name', 'assistants', '-id')

        # 筛选：车型
        model_filter = request.GET.get('model')
        if model_filter:
            queryset = queryset.filter(model__icontains=model_filter)

        # 筛选：VIN/零件编号
        vin = request.GET.get('vin_or_part_no')
        if vin:
            queryset = queryset.filter(vin_or_part_no__icontains=vin)

        # 筛选：试验名称
        test_name = request.GET.get('test_name')
        if test_name:
            queryset = queryset.filter(test_name__icontains=test_name)

        # 筛选：加入预警系统状态
        warning_status = request.GET.get('warning_system_status')
        if warning_status:
            queryset = queryset.filter(warning_system_status=warning_status)

        # 筛选：合同编号是否有内容（has_contract_no: true/false）
        has_contract_no = request.GET.get('has_contract_no')
        if has_contract_no is not None and has_contract_no != '':
            if has_contract_no.lower() in ['true', '1', 'yes']:
                # 有内容：非空且非null
                queryset = queryset.exclude(contract_no__isnull=True).exclude(contract_no__exact='')
            else:
                # 无内容：null或空字符串
                queryset = queryset.filter(Q(contract_no__isnull=True) | Q(contract_no__exact=''))

        # 筛选：任务提出人
        requester = request.GET.get('requester_name')
        if requester:
            queryset = queryset.filter(requester_name__icontains=requester)

        # 筛选：测试人员
        tester = request.GET.get('tester_name')
        if tester:
            queryset = queryset.filter(tester_name__icontains=tester)

        # 筛选：是否闭环
        is_closed = request.GET.get('is_closed')
        if is_closed is not None and is_closed != '':
            queryset = queryset.filter(is_closed=(is_closed.lower() in ['true', '1', 'yes']))

        # 筛选：样品状态（entry_exit_dispose_type）
        entry_exit_dispose_type = request.GET.get('entry_exit_dispose_type')
        if entry_exit_dispose_type is not None and entry_exit_dispose_type != '':
            if entry_exit_dispose_type == 'null' or entry_exit_dispose_type == '--':
                # 筛选null（未绑定进出登记或dispose_type为空）
                queryset = queryset.filter(
                    Q(entry_exit__isnull=True) | Q(entry_exit__dispose_type__isnull=True) | Q(entry_exit__dispose_type__exact='')
                )
            else:
                queryset = queryset.filter(entry_exit__dispose_type=entry_exit_dispose_type)

        # 筛选：出具报告（report_required: 是/否）
        report_required = request.GET.get('report_required')
        if report_required is not None and report_required != '':
            queryset = queryset.filter(report_required=report_required)

        # 筛选：时间范围（按日期维度，忽略时分秒）
        start_date = request.GET.get('schedule_start_from')
        end_date = request.GET.get('schedule_start_to')
        if start_date:
            # 解析日期字符串并转换为时区感知的 datetime
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                if settings.USE_TZ:
                    start_dt = timezone.make_aware(start_dt, timezone.get_current_timezone())
                queryset = queryset.filter(schedule_start__gte=start_dt)
            except ValueError:
                pass  # 日期格式错误时忽略该筛选条件
        if end_date:
            # 结束日期需要包含当天整天，所以用次日 00:00:00 作为上界（小于）
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                if settings.USE_TZ:
                    end_dt = timezone.make_aware(end_dt, timezone.get_current_timezone())
                queryset = queryset.filter(schedule_start__lt=end_dt)
            except ValueError:
                pass  # 日期格式错误时忽略该筛选条件

        # 分页
        page, page_size = get_pagination_params(request)
        items, total = paginate_queryset(queryset, page, page_size)

        serializer = MainRecordListSerializer(items, many=True)
        return Response.success(
            data={
                'items': serializer.data,
                'total': total,
                'page': page,
                'page_size': page_size,
            },
            message='获取任务列表成功'
        )

    # POST 创建
    serializer = MainRecordCreateUpdateSerializer(data=request.data)
    if serializer.is_valid():
        instance = serializer.save()
        # 刷新一次记录，保证'取消无样品'情况闭环
        services.refresh_main_closed(instance)
        output = MainRecordDetailSerializer(instance).data
        return Response.success(data=output, message='创建任务成功', status_code=201)
    # 区分校验错误：返回 validation_error 标识，前端据此显示"请填写完整信息"
    return Response.bad_request(message='请填写完整信息', data={'errors': serializer.errors, 'error_type': 'validation_error'})


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([AllowAny])
def main_record_detail(request, pk):
    """主记录详情 / 更新 / 删除"""
    try:
        instance = MainRecord.objects.select_related(
            'entry_exit', 'test_info', 'doc_approval'
        ).get(pk=pk)
    except MainRecord.DoesNotExist:
        return Response.not_found(message='任务不存在')

    if request.method == 'GET':
        serializer = MainRecordDetailSerializer(instance)
        return Response.success(data=serializer.data, message='获取任务详情成功')

    if request.method == 'PATCH':
        old_entry_exit_id = instance.entry_exit_id
        old_task_scenario = instance.task_scenario
        old_doc_requirement = instance.doc_requirement
        
        serializer = MainRecordCreateUpdateSerializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance = serializer.save()
            # 检查影响闭环的字段是否变化，触发闭环刷新
            if 'entry_exit_id' in request.data:
                new_entry_exit_id = instance.entry_exit_id
                if old_entry_exit_id != new_entry_exit_id:
                    services.refresh_main_closed(instance)
            # 任务场景或技术资料要求变化时也需要刷新闭环
            if ('task_scenario' in request.data and old_task_scenario != instance.task_scenario) or \
               ('doc_requirement' in request.data and old_doc_requirement != instance.doc_requirement):
                services.refresh_main_closed(instance)
            output = MainRecordDetailSerializer(instance).data
            return Response.success(data=output, message='更新任务成功')
        return Response.bad_request(message='更新任务失败', data=serializer.errors)

    if request.method == 'DELETE':
        instance.soft_delete()
        return Response.success(message='删除任务成功')


# ==================== EntryExit 视图 ====================

@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def entry_exit_list(request):
    """进出登记列表 / 创建"""
    if request.method == 'GET':
        queryset = EntryExit.objects.order_by('-created_at', '-id')

        # 筛选：处置类型 dispose_type
        dispose_type = request.GET.get('dispose_type')
        if dispose_type:
            queryset = queryset.filter(dispose_type=dispose_type)

        # 筛选：接收人
        receiver = request.GET.get('receiver_name')
        if receiver:
            queryset = queryset.filter(receiver_name__icontains=receiver)

        page, page_size = get_pagination_params(request)
        items, total = paginate_queryset(queryset, page, page_size)

        serializer = EntryExitSerializer(items, many=True)
        return Response.success(
            data={
                'items': serializer.data,
                'total': total,
                'page': page,
                'page_size': page_size,
            },
            message='获取进出登记列表成功'
        )

    # POST 创建
    serializer = EntryExitSerializer(data=request.data)
    if serializer.is_valid():
        instance = serializer.save()
        output = EntryExitSerializer(instance).data
        return Response.success(data=output, message='创建进出登记成功', status_code=201)
    return Response.bad_request(message='创建进出登记失败', data=serializer.errors)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([AllowAny])
def entry_exit_detail(request, pk):
    """进出登记详情 / 更新 / 删除"""
    try:
        instance = EntryExit.objects.get(pk=pk)
    except EntryExit.DoesNotExist:
        return Response.not_found(message='进出登记不存在')

    if request.method == 'GET':
        serializer = EntryExitSerializer(instance)
        return Response.success(data=serializer.data, message='获取进出登记详情成功')

    if request.method == 'PATCH':
        serializer = EntryExitSerializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance = serializer.save()
            output = EntryExitSerializer(instance).data
            return Response.success(data=output, message='更新进出登记成功')
        return Response.bad_request(message='更新进出登记失败', data=serializer.errors)

    if request.method == 'DELETE':
        try:
            instance.soft_delete()
            return Response.success(message='删除进出登记成功')
        except DjangoValidationError as e:
            return Response.bad_request(message=str(e.messages[0]) if e.messages else '删除失败')


@api_view(['POST'])
@permission_classes([AllowAny])
def entry_exit_submit(request, pk):
    """提交进出登记"""
    try:
        instance = EntryExit.objects.get(pk=pk)
    except EntryExit.DoesNotExist:
        return Response.not_found(message='进出登记不存在')

    try:
        services.submit_entry_exit(instance)
        output = EntryExitSerializer(instance).data
        return Response.success(data=output, message='提交成功')
    except DjangoValidationError as e:
        return Response.bad_request(message='提交失败', data=e.messages)


@api_view(['POST'])
@permission_classes([AllowAny])
def entry_exit_unsubmit(request, pk):
    """撤回进出登记"""
    try:
        instance = EntryExit.objects.get(pk=pk)
    except EntryExit.DoesNotExist:
        return Response.not_found(message='进出登记不存在')

    services.unsubmit_entry_exit(instance)
    output = EntryExitSerializer(instance).data
    return Response.success(data=output, message='撤回成功')


# ==================== TestInfo 视图 ====================

@api_view(['GET', 'PATCH'])
@permission_classes([AllowAny])
def test_info_by_main(request, main_id):
    """获取或创建主记录的试验信息（get_or_create）"""
    try:
        main = MainRecord.objects.get(pk=main_id)
    except MainRecord.DoesNotExist:
        return Response.not_found(message='主任务不存在')

    if request.method == 'GET':
        instance, created = TestInfo.objects.get_or_create(
            main=main,
            defaults={'status': STATUS_DRAFT}
        )
        serializer = TestInfoSerializer(instance)
        return Response.success(data=serializer.data, message='获取试验信息成功')

    if request.method == 'PATCH':
        instance, _created = TestInfo.objects.get_or_create(
            main=main,
            defaults={'status': STATUS_DRAFT}
        )

        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)

        move_teardown_from = ''
        teardown_url = data.get('teardown_attachment_url', '')
        if teardown_url and teardown_url.startswith(TEMP_UPLOAD_DIR):
            move_teardown_from = teardown_url
            data['teardown_attachment_url'] = get_final_upload_path(teardown_url, 'teardown_record')

        serializer = TestInfoSerializer(instance, data=data, partial=True)
        if not serializer.is_valid():
            return Response.bad_request(message='更新试验信息失败', data=serializer.errors)

        with transaction.atomic():
            instance = serializer.save()
            if move_teardown_from:
                transaction.on_commit(
                    lambda p=move_teardown_from: confirm_file_upload(p, 'teardown_record')
                )

        output = TestInfoSerializer(instance).data
        return Response.success(data=output, message='更新试验信息成功')


@api_view(['POST'])
@permission_classes([AllowAny])
def test_info_submit(request, pk):
    """提交试验信息"""
    try:
        instance = TestInfo.objects.get(pk=pk)
    except TestInfo.DoesNotExist:
        return Response.not_found(message='试验信息不存在')

    try:
        services.submit_test_info(instance)
        output = TestInfoSerializer(instance).data
        return Response.success(data=output, message='提交成功')
    except DjangoValidationError as e:
        return Response.bad_request(message='提交失败', data=e.messages)


@api_view(['POST'])
@permission_classes([AllowAny])
def test_info_unsubmit(request, pk):
    """撤回试验信息"""
    try:
        instance = TestInfo.objects.get(pk=pk)
    except TestInfo.DoesNotExist:
        return Response.not_found(message='试验信息不存在')

    services.unsubmit_test_info(instance)
    output = TestInfoSerializer(instance).data
    return Response.success(data=output, message='撤回成功')


# ==================== DocApproval 视图 ====================

@api_view(['GET', 'PATCH'])
@permission_classes([AllowAny])
def doc_approval_by_main(request, main_id):
    """获取或创建主记录的技术资料发放批准单（get_or_create）"""
    try:
        main = MainRecord.objects.get(pk=main_id)
    except MainRecord.DoesNotExist:
        return Response.not_found(message='主任务不存在')

    if request.method == 'GET':
        instance, created = DocApproval.objects.get_or_create(
            main=main,
            defaults={'status': STATUS_DRAFT}
        )
        serializer = DocApprovalSerializer(instance)
        return Response.success(data=serializer.data, message='获取技术资料批准单成功')

    if request.method == 'PATCH':
        instance, _created = DocApproval.objects.get_or_create(
            main=main,
            defaults={'status': STATUS_DRAFT}
        )

        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)

        move_doc_file_from = ''
        file_url = data.get('file_url', '')
        if file_url and file_url.startswith(TEMP_UPLOAD_DIR):
            move_doc_file_from = file_url
            data['file_url'] = get_final_upload_path(file_url, 'nvh_task_approval')

        serializer = DocApprovalSerializer(instance, data=data, partial=True)
        if not serializer.is_valid():
            return Response.bad_request(message='更新技术资料批准单失败', data=serializer.errors)

        with transaction.atomic():
            instance = serializer.save()
            if move_doc_file_from:
                transaction.on_commit(
                    lambda p=move_doc_file_from: confirm_file_upload(p, 'nvh_task_approval')
                )

        output = DocApprovalSerializer(instance).data
        return Response.success(data=output, message='更新技术资料批准单成功')


@api_view(['POST'])
@permission_classes([AllowAny])
def doc_approval_submit(request, pk):
    """提交技术资料发放批准单"""
    try:
        instance = DocApproval.objects.get(pk=pk)
    except DocApproval.DoesNotExist:
        return Response.not_found(message='技术资料批准单不存在')

    try:
        services.submit_doc_approval(instance)
        output = DocApprovalSerializer(instance).data
        return Response.success(data=output, message='提交成功')
    except DjangoValidationError as e:
        return Response.bad_request(message='提交失败', data=e.messages)


@api_view(['POST'])
@permission_classes([AllowAny])
def doc_approval_unsubmit(request, pk):
    """撤回技术资料发放批准单"""
    try:
        instance = DocApproval.objects.get(pk=pk)
    except DocApproval.DoesNotExist:
        return Response.not_found(message='技术资料批准单不存在')

    services.unsubmit_doc_approval(instance)
    output = DocApprovalSerializer(instance).data
    return Response.success(data=output, message='撤回成功')


# ==================== TestProcessAttachment 视图 ====================

@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def process_attachment_list(request):
    """过程记录附件列表 / 创建"""
    if request.method == 'GET':
        test_info_id = request.GET.get('test_info_id')
        if not test_info_id:
            return Response.bad_request(message='缺少 test_info_id 参数')

        queryset = TestProcessAttachment.objects.filter(
            test_info_id=test_info_id
        ).order_by('sort_no', 'id')

        serializer = TestProcessAttachmentSerializer(queryset, many=True)
        return Response.success(data=serializer.data, message='获取过程记录附件列表成功')

    # POST 创建
    data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)

    move_file_from = ''
    file_url = data.get('file_url', '')
    if file_url and file_url.startswith(TEMP_UPLOAD_DIR):
        move_file_from = file_url
        data['file_url'] = get_final_upload_path(file_url, 'nvh_test_process')

    serializer = TestProcessAttachmentSerializer(data=data)
    if not serializer.is_valid():
        return Response.bad_request(message='创建过程记录附件失败', data=serializer.errors)

    with transaction.atomic():
        instance = serializer.save()
        if move_file_from:
            transaction.on_commit(
                lambda p=move_file_from: confirm_file_upload(p, 'nvh_test_process')
            )

    output = TestProcessAttachmentSerializer(instance).data
    return Response.success(data=output, message='创建过程记录附件成功', status_code=201)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([AllowAny])
def process_attachment_detail(request, pk):
    """过程记录附件详情 / 更新 / 删除"""
    try:
        instance = TestProcessAttachment.objects.get(pk=pk)
    except TestProcessAttachment.DoesNotExist:
        return Response.not_found(message='过程记录附件不存在')

    if request.method == 'GET':
        serializer = TestProcessAttachmentSerializer(instance)
        return Response.success(data=serializer.data, message='获取过程记录附件详情成功')

    if request.method == 'PATCH':
        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)

        move_file_from = ''
        file_url = data.get('file_url', '')
        if file_url and file_url.startswith(TEMP_UPLOAD_DIR):
            move_file_from = file_url
            data['file_url'] = get_final_upload_path(file_url, 'nvh_test_process')

        serializer = TestProcessAttachmentSerializer(instance, data=data, partial=True)
        if not serializer.is_valid():
            return Response.bad_request(message='更新过程记录附件失败', data=serializer.errors)

        with transaction.atomic():
            instance = serializer.save()
            if move_file_from:
                transaction.on_commit(
                    lambda p=move_file_from: confirm_file_upload(p, 'nvh_test_process')
                )

        output = TestProcessAttachmentSerializer(instance).data
        return Response.success(data=output, message='更新过程记录附件成功')

    if request.method == 'DELETE':
        instance.soft_delete()
        return Response.success(message='删除过程记录附件成功')


# ==================== TestProcessList 视图 ====================

@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def process_list_options(request):
    """过程记录表名字典列表 / 新增"""
    if request.method == 'GET':
        queryset = TestProcessList.objects.order_by('id')
        serializer = TestProcessListSerializer(queryset, many=True)
        return Response.success(data=serializer.data, message='获取过程记录表名列表成功')

    # POST 创建（支持输入即新增）
    serializer = TestProcessListSerializer(data=request.data)
    if serializer.is_valid():
        instance = serializer.save()
        output = TestProcessListSerializer(instance).data
        return Response.success(data=output, message='创建过程记录表名成功', status_code=201)
    return Response.bad_request(message='创建过程记录表名失败', data=serializer.errors)


# ==================== 图片上传视图 ====================

@api_view(['POST'])
@permission_classes([AllowAny])
def upload_image(request):
    """
    通用图片上传接口
    文件先保存到临时目录，返回临时路径
    当表单保存成功时，由 confirm_upload 移动到最终目录
    """
    file = request.FILES.get('file')
    upload_type = request.POST.get('type', 'nvh_test_process')  # 默认过程记录

    if not file:
        return Response.bad_request(message='未选择文件')

    # 校验文件类型
    if file.content_type not in ALLOWED_IMAGE_TYPES:
        return Response.bad_request(message=f'不支持的文件类型: {file.content_type}，仅支持 jpg/png/webp/gif')

    # 校验文件大小
    if file.size > MAX_FILE_SIZE:
        return Response.bad_request(message=f'文件大小超过限制（最大 5MB）')

    # 生成唯一文件名
    ext = os.path.splitext(file.name)[1].lower()
    if not ext:
        ext = '.jpg'
    new_filename = f"{uuid.uuid4().hex}{ext}"

    # 保存到临时目录
    temp_dir = os.path.join(settings.MEDIA_ROOT, TEMP_UPLOAD_DIR)
    os.makedirs(temp_dir, exist_ok=True)
    temp_path = os.path.join(temp_dir, new_filename)

    try:
        with open(temp_path, 'wb+') as dest:
            for chunk in file.chunks():
                dest.write(chunk)
    except Exception as e:
        return Response.error(message=f'文件保存失败: {str(e)}')

    # 返回临时路径和目标类型
    temp_relative_path = f"{TEMP_UPLOAD_DIR}/{new_filename}"
    full_url = f"{settings.MEDIA_URL}{temp_relative_path}"

    return Response.success(
        data={
            'relative_path': temp_relative_path,
            'url': full_url,
            'filename': new_filename,
            'upload_type': upload_type,
            'is_temp': True,
        },
        message='上传成功（临时）'
    )


# ==================== 获取上次填写数据 ====================

@api_view(['GET'])
@permission_classes([AllowAny])
def get_last_main_record(request):
    """
    获取当前用户上一次成功创建的主记录数据，用于"使用上次填写"功能
    返回可回填的字段
    """
    # 获取最近一条创建的主记录（按创建时间倒序）
    last_record = MainRecord.objects.order_by('-created_at', '-id').first()
    
    if not last_record:
        return Response.success(data=None, message='暂无上次填写数据')
    
    # 只返回可回填的字段
    fillable_fields = {
        'model': last_record.model or '',
        'vin_or_part_no': last_record.vin_or_part_no or '',
        'test_name': last_record.test_name or '',
        'warning_system_status': last_record.warning_system_status or '',
        'requester_name': last_record.requester_name or '',
        'tester_name': last_record.tester_name or '',
        'assistants': last_record.assistants or '',
        'schedule_start': last_record.schedule_start.strftime('%Y-%m-%d') if last_record.schedule_start else None,
        'schedule_end': last_record.schedule_end.strftime('%Y-%m-%d') if last_record.schedule_end else None,
        'schedule_remark': last_record.schedule_remark or '',
        'test_location': last_record.test_location or '',
        'contract_no': last_record.contract_no or '',
    }
    
    return Response.success(data=fillable_fields, message='获取上次填写数据成功')


# ==================== 导出 Excel ====================

@api_view(['GET'])
@permission_classes([AllowAny])
def get_task_statistics(request):
    """
    获取任务统计数据
    返回：总任务数、本月任务数、本周任务数、本周已闭环数、本周未闭环数
    """
    now = timezone.localtime(timezone.now())

    # 获取本周的开始和结束时间（周一到周日）
    week_start = now - timedelta(days=now.weekday())  # 本周一
    week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
    week_end = week_start + timedelta(days=7)  # 下周一
    
    # 获取本月的开始和结束时间
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if now.month == 12:
        month_end = month_start.replace(year=now.year + 1, month=1)
    else:
        month_end = month_start.replace(month=now.month + 1)
    
    # 总任务数
    total_tasks = MainRecord.objects.count()
    
    # 本月任务数（排期开始时间在本月内）
    month_tasks = MainRecord.objects.filter(
        schedule_start__gte=month_start,
        schedule_start__lt=month_end
    ).count()
    
    # 本周任务数（排期开始时间在本周内）
    week_tasks = MainRecord.objects.filter(
        schedule_start__gte=week_start,
        schedule_start__lt=week_end
    ).count()
    
    # 本周已闭环任务数
    week_closed = MainRecord.objects.filter(
        schedule_start__gte=week_start,
        schedule_start__lt=week_end,
        is_closed=True
    ).count()
    
    # 本周未闭环任务数
    week_unclosed = MainRecord.objects.filter(
        schedule_start__gte=week_start,
        schedule_start__lt=week_end,
        is_closed=False
    ).count()
    
    return Response.success(
        data={
            'total_tasks': total_tasks,
            'month_tasks': month_tasks,
            'week_tasks': week_tasks,
            'week_closed': week_closed,
            'week_unclosed': week_unclosed
        },
        message='获取统计数据成功'
    )


@api_view(['GET'])
@permission_classes([AllowAny])
def export_main_records(request):
    """
    导出主记录列表为 Excel 文件
    支持与列表查询一致的筛选条件
    """
    if not HAS_OPENPYXL:
        return Response.bad_request(message='服务器未安装 openpyxl 库，无法导出 Excel')
    
    # 复用列表查询的筛选逻辑
    queryset = MainRecord.objects.select_related(
        'entry_exit', 'test_info', 'doc_approval'
    ).order_by('tester_name', '-schedule_start', '-id')

    # 筛选：车型
    model_filter = request.GET.get('model')
    if model_filter:
        queryset = queryset.filter(model__icontains=model_filter)

    # 筛选：VIN/零件编号
    vin = request.GET.get('vin_or_part_no')
    if vin:
        queryset = queryset.filter(vin_or_part_no__icontains=vin)

    # 筛选：试验名称
    test_name = request.GET.get('test_name')
    if test_name:
        queryset = queryset.filter(test_name__icontains=test_name)

    # 筛选：加入预警系统状态
    warning_status = request.GET.get('warning_system_status')
    if warning_status:
        queryset = queryset.filter(warning_system_status=warning_status)

    # 筛选：合同编号是否有内容
    has_contract_no = request.GET.get('has_contract_no')
    if has_contract_no is not None and has_contract_no != '':
        if has_contract_no.lower() in ['true', '1', 'yes']:
            queryset = queryset.exclude(contract_no__isnull=True).exclude(contract_no__exact='')
        else:
            queryset = queryset.filter(Q(contract_no__isnull=True) | Q(contract_no__exact=''))

    # 筛选：任务提出人
    requester = request.GET.get('requester_name')
    if requester:
        queryset = queryset.filter(requester_name__icontains=requester)

    # 筛选：测试人员
    tester = request.GET.get('tester_name')
    if tester:
        queryset = queryset.filter(tester_name__icontains=tester)

    # 筛选：是否闭环
    is_closed = request.GET.get('is_closed')
    if is_closed is not None and is_closed != '':
        queryset = queryset.filter(is_closed=(is_closed.lower() in ['true', '1', 'yes']))

    # 筛选：样品状态
    entry_exit_dispose_type = request.GET.get('entry_exit_dispose_type')
    if entry_exit_dispose_type is not None and entry_exit_dispose_type != '':
        if entry_exit_dispose_type == 'null' or entry_exit_dispose_type == '--':
            queryset = queryset.filter(
                Q(entry_exit__isnull=True) | Q(entry_exit__dispose_type__isnull=True) | Q(entry_exit__dispose_type__exact='')
            )
        else:
            queryset = queryset.filter(entry_exit__dispose_type=entry_exit_dispose_type)

    # 筛选：出具报告
    report_required = request.GET.get('report_required')
    if report_required is not None and report_required != '':
        queryset = queryset.filter(report_required=report_required)

    # 筛选：时间范围
    start_date = request.GET.get('schedule_start_from')
    end_date = request.GET.get('schedule_start_to')
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            if settings.USE_TZ:
                start_dt = timezone.make_aware(start_dt, timezone.get_current_timezone())
            queryset = queryset.filter(schedule_start__gte=start_dt)
        except ValueError:
            pass
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            if settings.USE_TZ:
                end_dt = timezone.make_aware(end_dt, timezone.get_current_timezone())
            queryset = queryset.filter(schedule_start__lt=end_dt)
        except ValueError:
            pass

    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '试验任务单'

    # 定义表头（与前端表格列顺序一致）
    headers = [
        '闭环状态', '样品状态', '预警', '车型', 'VIN/零件编号', '试验名称',
        '测试人员', '协助人员', '提出人', '排期开始', '排期结束', '排期备注',
        '地点', '报告', '合同编号', '备注'
    ]

    # 设置表头样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = openpyxl.styles.PatternFill(start_color='409EFF', end_color='409EFF', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, record in enumerate(queryset, 2):
        # 闭环状态
        ws.cell(row=row_idx, column=1, value='已闭环' if record.is_closed else '未闭环')
        # 样品状态
        dispose_type = ''
        if record.entry_exit and not record.entry_exit.is_deleted:
            dispose_type = record.entry_exit.dispose_type or '--'
        else:
            dispose_type = '--'
        ws.cell(row=row_idx, column=2, value=dispose_type)
        # 预警
        ws.cell(row=row_idx, column=3, value=record.warning_system_status or '')
        # 车型
        ws.cell(row=row_idx, column=4, value=record.model or '')
        # VIN/零件编号
        ws.cell(row=row_idx, column=5, value=record.vin_or_part_no or '')
        # 试验名称
        ws.cell(row=row_idx, column=6, value=record.test_name or '')
        # 测试人员
        ws.cell(row=row_idx, column=7, value=record.tester_name or '')
        # 协助人员
        ws.cell(row=row_idx, column=8, value=record.assistants or '--')
        # 提出人
        ws.cell(row=row_idx, column=9, value=record.requester_name or '')
        # 排期开始
        schedule_start = record.schedule_start.strftime('%Y-%m-%d') if record.schedule_start else '--'
        ws.cell(row=row_idx, column=10, value=schedule_start)
        # 排期结束
        schedule_end = record.schedule_end.strftime('%Y-%m-%d') if record.schedule_end else '--'
        ws.cell(row=row_idx, column=11, value=schedule_end)
        # 排期备注
        ws.cell(row=row_idx, column=12, value=record.schedule_remark or '--')
        # 地点
        ws.cell(row=row_idx, column=13, value=record.test_location or '')
        # 报告
        ws.cell(row=row_idx, column=14, value=record.report_required or '')
        # 合同编号
        ws.cell(row=row_idx, column=15, value=record.contract_no or '否')
        # 备注
        ws.cell(row=row_idx, column=16, value=record.remark or '--')

        # 设置数据行样式
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # 调整列宽
    column_widths = [10, 10, 10, 12, 20, 30, 12, 12, 12, 12, 12, 15, 15, 8, 12, 15]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # 生成文件名
    today = datetime.now().strftime('%Y%m%d')
    filename = f'试验任务单_{today}.xlsx'

    # 写入响应
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # 使用 RFC 5987 编码处理中文文件名
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{filename}"
    
    return response
