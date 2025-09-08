#!/usr/bin/env python3
"""
简化版MySQL数据导出到Excel工具
专门用于导出悬架隔振率测试数据示例
"""

import os
import sys
import pymysql
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# 检查依赖包
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("❌ 缺少依赖包 openpyxl")
    print("💡 请运行: pip install openpyxl")
    sys.exit(1)

# 添加Django项目路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'nvh_backend.settings')

# 导入Django配置
import django
django.setup()

from django.conf import settings


def get_db_connection():
    """获取数据库连接"""
    try:
        db_config = settings.DATABASES['default']
        connection = pymysql.connect(
            host=db_config['HOST'],
            port=int(db_config['PORT']),
            user=db_config['USER'],
            password=db_config['PASSWORD'],
            database=db_config['NAME'],
            charset='utf8mb4'
        )
        print(f"✅ 成功连接到数据库: {db_config['NAME']}")
        return connection
    except Exception as e:
        print(f"❌ 数据库连接失败: {str(e)}")
        return None


def export_table_to_excel(table_name, field_mappings=None):
    """
    导出指定表到Excel文件
    
    Args:
        table_name: 要导出的表名
        field_mappings: 字段中英文映射字典
    """
    # 默认字段映射
    if field_mappings is None:
        field_mappings = {
            "id": "序号",
            "test_date": "测试日期",
            "test_location": "测试地点",
            "test_engineer": "测试工程师",
            "suspension_type": "悬架类型",
            "tire_pressure": "轮胎气压",
            "test_condition": "测试工况",
            "vehicle_model_id": "车型编号"
        }
    
    connection = get_db_connection()
    if not connection:
        return False
    
    try:
        print(f"\\n🔄 开始导出表: {table_name}")
        
        # 获取表的列信息
        with connection.cursor() as cursor:
            cursor.execute(f"DESCRIBE `{table_name}`")
            columns = [column[0] for column in cursor.fetchall()]
        
        if not columns:
            print(f"❌ 表 {table_name} 没有列信息")
            return False
        
        # 查询所有数据
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT * FROM `{table_name}`")
            rows = cursor.fetchall()
        
        print(f"📊 表 {table_name} 共有 {len(rows)} 条数据")
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = table_name
        
        # 设置标题行样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        chinese_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # 第一行：英文字段名
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = header_font
            cell.fill = header_fill
        
        # 第二行：中文字段说明
        for col_idx, column in enumerate(columns, 1):
            chinese_name = field_mappings.get(column, column)
            cell = ws.cell(row=2, column=col_idx, value=chinese_name)
            cell.fill = chinese_fill
            cell.font = Font(bold=True)
        
        # 第三行开始：实际数据
        for row_idx, row in enumerate(rows, 3):
            for col_idx, value in enumerate(row, 1):
                # 处理特殊数据类型
                if value is None:
                    display_value = ""
                elif isinstance(value, datetime):
                    display_value = value.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    display_value = str(value)
                
                ws.cell(row=row_idx, column=col_idx, value=display_value)
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # 设置列宽，最小10，最大50
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存Excel文件
        output_filename = f"{table_name}.xlsx"
        wb.save(output_filename)
        
        print(f"✅ 成功导出 {table_name} 到 {output_filename}")
        print(f"📁 文件保存位置: {os.path.abspath(output_filename)}")
        return True
        
    except Exception as e:
        print(f"❌ 导出表 {table_name} 失败: {str(e)}")
        return False
    
    finally:
        connection.close()


def get_all_tables():
    """获取所有表列表"""
    connection = get_db_connection()
    if not connection:
        return []
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("SHOW TABLES")
            all_tables = [table[0] for table in cursor.fetchall()]
            
            # 排除Django系统表
            excluded_tables = {
                'auth_group', 'auth_group_permissions', 'auth_permission',
                'auth_user', 'auth_user_groups', 'auth_user_user_permissions',
                'django_admin_log', 'django_content_type', 'django_migrations',
                'django_session', 'oidc_rp_oidcuser'
            }
            
            user_tables = [table for table in all_tables if table not in excluded_tables]
            return user_tables
    
    except Exception as e:
        print(f"❌ 获取表列表失败: {str(e)}")
        return []
    
    finally:
        connection.close()


def main():
    """主函数"""
    print("=" * 60)
    print("🚀 MySQL数据导出到Excel工具")
    print("=" * 60)
    
    # 示例：导出悬架隔振率测试表
    table_name = "vehicle_suspension_isolation_tests"
    
    # 悬架隔振率测试字段映射
    field_mappings = {
        "id": "序号",
        "test_date": "测试日期",
        "test_location": "测试地点",
        "test_engineer": "测试工程师",
        "suspension_type": "悬架类型",
        "tire_pressure": "轮胎气压",
        "test_condition": "测试工况",
        "vehicle_model_id": "车型编号"
    }
    
    print("\\n📋 可选操作:")
    print("1. 导出悬架隔振率测试表")
    print("2. 查看所有表列表")
    print("3. 导出指定表")
    
    choice = input("\\n请选择操作 (1-3): ").strip()
    
    if choice == "1":
        export_table_to_excel(table_name, field_mappings)
    
    elif choice == "2":
        tables = get_all_tables()
        print(f"\\n📋 数据库中的用户表 (共{len(tables)}个):")
        for i, table in enumerate(tables, 1):
            print(f"   {i:2d}. {table}")
    
    elif choice == "3":
        tables = get_all_tables()
        print(f"\\n📋 可用表列表:")
        for i, table in enumerate(tables, 1):
            print(f"   {i:2d}. {table}")
        
        table_input = input("\\n请输入要导出的表名: ").strip()
        if table_input in tables:
            export_table_to_excel(table_input)
        else:
            print(f"❌ 表 '{table_input}' 不存在")
    
    else:
        print("❌ 无效选择")


if __name__ == "__main__":
    main()