#!/usr/bin/env python3
"""
MySQL数据导出到Excel工具
用于导出数据库中的所有用户表数据到Excel文件
"""

import os
import sys
import pymysql
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# 尝试导入openpyxl，如果没有安装则提示安装
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("❌ 缺少依赖包 openpyxl")
    print("💡 请运行: pip install openpyxl")
    sys.exit(1)

# 添加Django项目路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'nvh_backend.settings')

# 导入Django配置
import django
django.setup()

from django.conf import settings


class MySQLToExcelExporter:
    """MySQL数据导出到Excel的工具类"""
    
    def __init__(self):
        """初始化数据库连接配置"""
        db_config = settings.DATABASES['default']
        self.db_config = {
            'host': db_config['HOST'],
            'port': int(db_config['PORT']),
            'user': db_config['USER'],
            'password': db_config['PASSWORD'],
            'database': db_config['NAME'],
            'charset': 'utf8mb4'
        }
        
        # 中英文字段映射 - 悬架隔振率测试示例
        self.field_mappings = {
            "id": "序号",
            "test_date": "测试日期",
            "test_location": "测试地点", 
            "test_engineer": "测试工程师",
            "suspension_type": "悬架类型",
            "tire_pressure": "轮胎气压",
            "test_condition": "测试工况",
            "vehicle_model_id": "车型编号",
            "created_at": "创建时间",
            "updated_at": "更新时间",
            "measuring_point": "测点名称",
            "x_active_value": "X方向主动端",
            "x_passive_value": "X方向被动端", 
            "x_isolation_rate": "X方向隔振率",
            "y_active_value": "Y方向主动端",
            "y_passive_value": "Y方向被动端",
            "y_isolation_rate": "Y方向隔振率",
            "z_active_value": "Z方向主动端",
            "z_passive_value": "Z方向被动端",
            "z_isolation_rate": "Z方向隔振率",
            "layout_image_path": "测试布置图路径",
            "curve_image_path": "测试数据曲线图路径"
        }
        
        # 需要排除的Django系统表
        self.excluded_tables = {
            'auth_group', 'auth_group_permissions', 'auth_permission',
            'auth_user', 'auth_user_groups', 'auth_user_user_permissions',
            'django_admin_log', 'django_content_type', 'django_migrations',
            'django_session', 'oidc_rp_oidcuser'
        }
    
    def get_connection(self):
        """获取数据库连接"""
        try:
            connection = pymysql.connect(**self.db_config)
            print(f"✅ 成功连接到数据库: {self.db_config['database']}")
            return connection
        except Exception as e:
            print(f"❌ 数据库连接失败: {str(e)}")
            return None
    
    def get_all_user_tables(self, connection):
        """获取所有用户表（排除Django系统表）"""
        try:
            with connection.cursor() as cursor:
                cursor.execute("SHOW TABLES")
                all_tables = [table[0] for table in cursor.fetchall()]
                
                # 过滤掉Django系统表
                user_tables = [table for table in all_tables 
                              if table not in self.excluded_tables]
                
                print(f"📋 发现 {len(user_tables)} 个用户表:")
                for table in user_tables:
                    print(f"   - {table}")
                
                return user_tables
        except Exception as e:
            print(f"❌ 获取表列表失败: {str(e)}")
            return []
    
    def get_table_columns(self, connection, table_name):
        """获取表的列信息"""
        try:
            with connection.cursor() as cursor:
                cursor.execute(f"DESCRIBE `{table_name}`")
                columns = [column[0] for column in cursor.fetchall()]
                return columns
        except Exception as e:
            print(f"❌ 获取表 {table_name} 列信息失败: {str(e)}")
            return []
    
    def get_table_column_info(self, connection, table_name):
        """获取表的详细列信息，包括是否允许NULL"""
        try:
            with connection.cursor() as cursor:
                cursor.execute(f"DESCRIBE `{table_name}`")
                column_info = cursor.fetchall()
                # 返回格式: [(field_name, type, null, key, default, extra), ...]
                # null字段: 'YES'表示允许NULL, 'NO'表示NOT NULL
                return column_info
        except Exception as e:
            print(f"❌ 获取表 {table_name} 详细列信息失败: {str(e)}")
            return []
    
    def get_chinese_column_name(self, english_name, is_not_null=False):
        """获取字段的中文名称，并标记是否为NOT NULL"""
        chinese_name = self.field_mappings.get(english_name, english_name)
        if is_not_null:
            chinese_name += " (*必填)"
        return chinese_name
    
    def export_table_to_excel(self, table_name):
        """导出指定表到Excel文件"""
        connection = self.get_connection()
        if not connection:
            return False
        
        try:
            print(f"\n🔄 开始导出表: {table_name}")
            
            # 获取表的详细列信息
            column_info = self.get_table_column_info(connection, table_name)
            if not column_info:
                print(f"❌ 表 {table_name} 没有列信息")
                return False
            
            # 提取列名和NOT NULL信息
            columns = [col[0] for col in column_info]  # 字段名
            not_null_flags = [col[2] == 'NO' for col in column_info]  # NULL字段为'NO'表示NOT NULL
            
            # 查询所有数据
            with connection.cursor() as cursor:
                cursor.execute(f"SELECT * FROM `{table_name}`")
                rows = cursor.fetchall()
            
            print(f"📊 表 {table_name} 共有 {len(rows)} 条数据")
            
            # 显示NOT NULL字段信息
            not_null_fields = [columns[i] for i, is_not_null in enumerate(not_null_flags) if is_not_null]
            if not_null_fields:
                print(f"📝 NOT NULL字段: {', '.join(not_null_fields)}")
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = table_name
            
            # 设置标题行样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            chinese_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            # NOT NULL字段特殊样式（红色背景）
            not_null_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            not_null_font = Font(bold=True, color="CC0000")
            
            # 第一行：英文字段名
            for col_idx, column in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=column)
                cell.font = header_font
                cell.fill = header_fill
            
            # 第二行：中文字段说明（标记NOT NULL字段）
            for col_idx, (column, is_not_null) in enumerate(zip(columns, not_null_flags), 1):
                chinese_name = self.get_chinese_column_name(column, is_not_null)
                cell = ws.cell(row=2, column=col_idx, value=chinese_name)
                
                if is_not_null:
                    cell.fill = not_null_fill
                    cell.font = not_null_font
                else:
                    cell.fill = chinese_fill
                    cell.font = Font(bold=True)
            
            # 第三行开始：实际数据
            for row_idx, row in enumerate(rows, 3):
                for col_idx, value in enumerate(row, 1):
                    # 处理特殊数据类型
                    if value is None:
                        display_value = ""
                    elif isinstance(value, datetime):
                        display_value = value.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        display_value = str(value)
                    
                    ws.cell(row=row_idx, column=col_idx, value=display_value)
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                # 设置列宽，最小10，最大50
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存Excel文件
            output_filename = f"{table_name}.xlsx"
            wb.save(output_filename)
            
            print(f"✅ 成功导出 {table_name} 到 {output_filename}")
            print(f"📁 文件保存位置: {os.path.abspath(output_filename)}")
            if not_null_fields:
                print(f"📝 已标记 {len(not_null_fields)} 个NOT NULL字段（带*必填标记）")
            return True
            
        except Exception as e:
            print(f"❌ 导出表 {table_name} 失败: {str(e)}")
            return False
        
        finally:
            connection.close()
    
    def export_all_tables(self):
        """导出所有用户表到Excel文件"""
        connection = self.get_connection()
        if not connection:
            return
        
        try:
            # 获取所有用户表
            user_tables = self.get_all_user_tables(connection)
            if not user_tables:
                print("❌ 没有发现用户表")
                return
            
            # 创建导出目录
            export_dir = f"excel_exports_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            os.makedirs(export_dir, exist_ok=True)
            print(f"📁 创建导出目录: {export_dir}")
            
            success_count = 0
            
            # 逐个导出表
            for table_name in user_tables:
                print(f"\n🔄 导出表: {table_name}")
                
                try:
                    # 获取表的详细列信息
                    column_info = self.get_table_column_info(connection, table_name)
                    if not column_info:
                        print(f"⚠️  跳过表 {table_name}: 无列信息")
                        continue
                    
                    # 提取列名和NOT NULL信息
                    columns = [col[0] for col in column_info]
                    not_null_flags = [col[2] == 'NO' for col in column_info]
                    
                    # 查询所有数据
                    with connection.cursor() as cursor:
                        cursor.execute(f"SELECT * FROM `{table_name}`")
                        rows = cursor.fetchall()
                    
                    print(f"📊 表 {table_name} 共有 {len(rows)} 条数据")
                    
                    # 显示NOT NULL字段
                    not_null_fields = [columns[i] for i, is_not_null in enumerate(not_null_flags) if is_not_null]
                    if not_null_fields:
                        print(f"📝 NOT NULL字段: {', '.join(not_null_fields)}")
                    
                    # 创建Excel工作簿
                    wb = Workbook()
                    ws = wb.active
                    ws.title = table_name
                    
                    # 设置标题行样式
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    chinese_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    # NOT NULL字段特殊样式
                    not_null_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    not_null_font = Font(bold=True, color="CC0000")
                    
                    # 第一行：英文字段名
                    for col_idx, column in enumerate(columns, 1):
                        cell = ws.cell(row=1, column=col_idx, value=column)
                        cell.font = header_font
                        cell.fill = header_fill
                    
                    # 第二行：中文字段说明（标记NOT NULL字段）
                    for col_idx, (column, is_not_null) in enumerate(zip(columns, not_null_flags), 1):
                        chinese_name = self.get_chinese_column_name(column, is_not_null)
                        cell = ws.cell(row=2, column=col_idx, value=chinese_name)
                        
                        if is_not_null:
                            cell.fill = not_null_fill
                            cell.font = not_null_font
                        else:
                            cell.fill = chinese_fill
                            cell.font = Font(bold=True)
                    
                    # 第三行开始：实际数据
                    for row_idx, row in enumerate(rows, 3):
                        for col_idx, value in enumerate(row, 1):
                            # 处理特殊数据类型
                            if value is None:
                                display_value = ""
                            elif isinstance(value, datetime):
                                display_value = value.strftime("%Y-%m-%d %H:%M:%S")
                            else:
                                display_value = str(value)
                            
                            ws.cell(row=row_idx, column=col_idx, value=display_value)
                    
                    # 自动调整列宽
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        
                        # 设置列宽，最小10，最大50
                        adjusted_width = min(max(max_length + 2, 10), 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # 保存Excel文件
                    output_filename = os.path.join(export_dir, f"{table_name}.xlsx")
                    wb.save(output_filename)
                    
                    print(f"✅ 成功导出 {table_name}")
                    if not_null_fields:
                        print(f"📝 已标记 {len(not_null_fields)} 个NOT NULL字段")
                    success_count += 1
                    
                except Exception as e:
                    print(f"❌ 导出表 {table_name} 失败: {str(e)}")
                    continue
            
            print(f"\n🎉 导出完成!")
            print(f"📊 总计: {len(user_tables)} 个表，成功: {success_count} 个")
            print(f"📁 文件保存目录: {os.path.abspath(export_dir)}")
            
        except Exception as e:
            print(f"❌ 批量导出失败: {str(e)}")
        
        finally:
            connection.close()
    
    def export_single_table(self, table_name):
        """导出单个指定表"""
        connection = self.get_connection()
        if not connection:
            return False
        
        try:
            # 检查表是否存在
            with connection.cursor() as cursor:
                cursor.execute("SHOW TABLES")
                all_tables = [table[0] for table in cursor.fetchall()]
                
                if table_name not in all_tables:
                    print(f"❌ 表 {table_name} 不存在")
                    print(f"📋 可用的表: {', '.join(all_tables)}")
                    return False
            
            # 导出表
            result = self.export_table_to_excel(table_name)
            return result
            
        except Exception as e:
            print(f"❌ 导出失败: {str(e)}")
            return False
        
        finally:
            connection.close()


def main():
    """主函数"""
    print("=" * 60)
    print("🚀 MySQL数据导出到Excel工具")
    print("=" * 60)
    
    exporter = MySQLToExcelExporter()
    
    print("\n请选择操作:")
    print("1. 导出所有用户表")
    print("2. 导出指定表")
    print("3. 查看所有表列表")
    
    choice = input("\n请输入选择 (1-3): ").strip()
    
    if choice == "1":
        print("\n开始导出所有用户表...")
        exporter.export_all_tables()
    
    elif choice == "2":
        table_name = input("\n请输入要导出的表名: ").strip()
        if table_name:
            exporter.export_single_table(table_name)
        else:
            print("❌ 请输入有效的表名")
    
    elif choice == "3":
        connection = exporter.get_connection()
        if connection:
            try:
                user_tables = exporter.get_all_user_tables(connection)
                print(f"\n📋 共有 {len(user_tables)} 个用户表可供导出:")
                for i, table in enumerate(user_tables, 1):
                    print(f"   {i:2d}. {table}")
            finally:
                connection.close()
    
    else:
        print("❌ 无效的选择")


if __name__ == "__main__":
    main()